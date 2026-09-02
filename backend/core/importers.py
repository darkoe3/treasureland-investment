import hashlib
import math
import zipfile
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

import openpyxl
from django.core.exceptions import ValidationError
from django.utils.text import get_valid_filename
from openpyxl.utils import get_column_letter

from .models import DailySheet, DailySheetGame, TPMCode, WeeklyGameSchedule, money


MAX_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_SHEETS = 10
MAX_ZIP_ENTRIES = 200
MAX_UNCOMPRESSED_BYTES = 30 * 1024 * 1024
MAX_INSPECTED_COLUMNS = 20
TRANSACTION_ROWS = range(5, 225)
SALES_COLUMNS = range(3, 10)
REQUIRED_SHEETS = [
    "ENTER GAME DATA HERE",
    "REGISTER SUB-AGENT",
    "MUSA RESULTS",
    "Premier Games",
    "Sheet2",
]
UNSUPPORTED_PARTS = (
    "vbaProject.bin",
    "xl/externalLinks/",
    "xl/embeddings/",
    "xl/activeX/",
)
UNSUPPORTED_EXTENSIONS = (".exe", ".dll", ".js", ".jar", ".bat", ".cmd", ".ps1", ".scr", ".com")

GAME_ALIASES = {
    "f/chance": "Fairchance",
    "fairchance": "Fairchance",
    "diamondq": "Diamond",
    "diamond": "Diamond",
    "inter": "International",
    "international": "International",
    "fortune": "Fortune",
    "mk ii": "Mark II",
    "mark ii": "Mark II",
    "c/master": "Club Master",
    "clubmaster": "Club Master",
    "club master": "Club Master",
    "o6": "06",
    "06": "06",
    "msp": "Monday Special",
    "monday special": "Monday Special",
}


@dataclass
class WorkbookParseResult:
    payload: dict
    warnings: list
    errors: list


def safe_filename(name):
    return get_valid_filename(Path(name or "upload.xlsx").name)[:255]


def excel_ref(row, column):
    return f"{get_column_letter(column)}{row}"


def normalize_identifier(value, cell_ref, warnings):
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        warnings.append({"cell": cell_ref, "message": "Numeric identifier may have lost leading zeroes."})
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            warnings.append({"cell": cell_ref, "message": "Numeric identifier may have lost leading zeroes."})
            return str(int(value))
        warnings.append({"cell": cell_ref, "message": "Identifier contains decimal digits."})
        return format(value, "f").rstrip("0").rstrip(".")
    return str(value).strip()


def normalize_game_name(value):
    text = str(value or "").strip()
    if not text:
        return ""
    return GAME_ALIASES.get(text.lower(), text)


def validate_workbook_bytes(uploaded_file):
    filename = safe_filename(uploaded_file.name)
    if not filename.lower().endswith(".xlsx"):
        raise ValidationError("Only .xlsx files are supported.")
    if uploaded_file.size and uploaded_file.size > MAX_UPLOAD_BYTES:
        raise ValidationError("Workbook is larger than the 5 MB upload limit.")
    data = uploaded_file.read()
    uploaded_file.seek(0)
    if len(data) > MAX_UPLOAD_BYTES:
        raise ValidationError("Workbook is larger than the 5 MB upload limit.")
    if not data.startswith(b"PK\x03\x04"):
        raise ValidationError("File content is not a valid .xlsx package.")
    try:
        with zipfile.ZipFile(BytesIO(data)) as package:
            infos = package.infolist()
            if len(infos) > MAX_ZIP_ENTRIES:
                raise ValidationError("Workbook package contains too many files.")
            if sum(info.file_size for info in infos) > MAX_UNCOMPRESSED_BYTES:
                raise ValidationError("Workbook package is too large when decompressed.")
            names = package.namelist()
            if any(part in name for part in UNSUPPORTED_PARTS for name in names):
                raise ValidationError("Workbook contains macros, external links or embedded unsupported content.")
            if any(name.lower().endswith(UNSUPPORTED_EXTENSIONS) for name in names):
                raise ValidationError("Workbook contains embedded unsupported executable content.")
            for name in names:
                if name.endswith(".rels"):
                    rels = package.read(name)[:1024 * 1024]
                    if b'TargetMode="External"' in rels or b"TargetMode='External'" in rels:
                        raise ValidationError("Workbook contains external links or references.")
    except zipfile.BadZipFile as exc:
        raise ValidationError("Workbook is corrupt or unsupported.") from exc
    return data


def load_workbook(data):
    try:
        workbook = openpyxl.load_workbook(BytesIO(data), data_only=False, read_only=True, keep_links=False)
    except Exception as exc:
        raise ValidationError("Workbook is password-protected, corrupt or unsupported.") from exc
    if len(workbook.sheetnames) > MAX_SHEETS:
        raise ValidationError("Workbook has too many worksheets.")
    missing = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
    if missing:
        raise ValidationError(f"Workbook is missing required worksheet(s): {', '.join(missing)}.")
    return workbook


def parse_workbook_date(raw_value):
    if isinstance(raw_value, datetime):
        return raw_value.date()
    if hasattr(raw_value, "date"):
        return raw_value.date()
    if isinstance(raw_value, str) and raw_value.strip():
        try:
            return datetime.fromisoformat(raw_value.strip()).date()
        except ValueError:
            return None
    return None


def decimal_from_cell(cell, cell_ref, errors):
    value = cell.value
    if value in (None, ""):
        return money("0")
    if getattr(cell, "data_type", None) == "f":
        errors.append({"cell": cell_ref, "message": "Formula cells are not allowed in sales rows."})
        return money("0")
    if isinstance(value, bool) or isinstance(value, str):
        errors.append({"cell": cell_ref, "message": "Sales amount must be a literal number."})
        return money("0")
    if isinstance(value, float) and not math.isfinite(value):
        errors.append({"cell": cell_ref, "message": "Sales amount must be finite."})
        return money("0")
    try:
        amount = money(Decimal(str(value)))
    except (InvalidOperation, ValueError):
        errors.append({"cell": cell_ref, "message": "Sales amount is invalid."})
        return money("0")
    if amount < 0:
        errors.append({"cell": cell_ref, "message": "Sales amount cannot be negative."})
    if amount > Decimal("999999999999.99"):
        errors.append({"cell": cell_ref, "message": "Sales amount is too large."})
    return amount


def registration_lookup(sheet, warnings, errors):
    mapping = {}
    reverse = {}
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, max_row=min(sheet.max_row, 300), min_col=2, max_col=4), start=2):
        sub_cell, terminal_cell, name_cell = row
        sub_ref = excel_ref(row_number, 2)
        terminal_ref = excel_ref(row_number, 3)
        sub_code = normalize_identifier(sub_cell.value, sub_ref, warnings)
        terminal = normalize_identifier(terminal_cell.value, terminal_ref, warnings)
        workbook_name = str(name_cell.value or "").strip()
        if not sub_code and not terminal:
            continue
        if not sub_code or not terminal:
            errors.append({"cell": sub_ref, "message": "Registration row requires both SUB AGT NOS and TERMINAL NOS."})
            continue
        if sub_code in mapping and mapping[sub_code]["terminal"] != terminal:
            errors.append({"cell": sub_ref, "message": "SUB AGT NOS maps to multiple TERMINAL NOS values."})
        mapping[sub_code] = {"terminal": terminal, "workbook_name": workbook_name}
        reverse.setdefault(terminal.lower(), set()).add(sub_code)
    for terminal, sub_codes in reverse.items():
        if len(sub_codes) > 1:
            errors.append({"cell": "REGISTER SUB-AGENT", "message": f"TERMINAL NOS {terminal} is mapped from multiple SUB AGT NOS values."})
    return mapping


def scheduled_games_for_date(transaction_date, existing_sheet):
    if existing_sheet:
        rows = existing_sheet.sheet_games.select_related("game").order_by("display_order", "id")
        return {row.game_name_snapshot.lower(): row for row in rows}
    schedules = (
        WeeklyGameSchedule.objects.select_related("game")
        .filter(weekday=transaction_date.isoweekday(), is_active=True, game__is_active=True)
        .order_by("display_order", "id")
    )
    return {schedule.game.name.lower(): schedule for schedule in schedules}


def schedule_snapshot_for_date(transaction_date, existing_sheet):
    if existing_sheet:
        rows = existing_sheet.sheet_games.order_by("display_order", "id")
        return [
            {
                "game_name": row.game_name_snapshot,
                "is_whole_day": row.is_whole_day_snapshot,
                "closing_time": row.closing_time_snapshot.isoformat() if row.closing_time_snapshot else None,
                "draw_time": row.draw_time_snapshot.isoformat() if row.draw_time_snapshot else None,
                "display_order": row.display_order,
            }
            for row in rows
        ]
    schedules = (
        WeeklyGameSchedule.objects.select_related("game")
        .filter(weekday=transaction_date.isoweekday(), is_active=True, game__is_active=True)
        .order_by("display_order", "id")
    )
    return [
        {
            "game_name": schedule.game.name,
            "is_whole_day": schedule.is_whole_day,
            "closing_time": schedule.closing_time.isoformat() if schedule.closing_time else None,
            "draw_time": schedule.draw_time.isoformat() if schedule.draw_time else None,
            "display_order": schedule.display_order,
        }
        for schedule in schedules
    ]


def build_game_columns(raw_sheet, scheduled_by_name, errors):
    columns = []
    seen = {}
    for col in SALES_COLUMNS:
        header = normalize_game_name(raw_sheet.cell(3, col).value)
        letter = get_column_letter(col)
        if not header:
            has_sales = any(raw_sheet.cell(row, col).value not in (None, "", 0) for row in TRANSACTION_ROWS)
            if has_sales:
                errors.append({"cell": f"{letter}3", "message": "Sales exist under a blank game header."})
            continue
        scheduled = scheduled_by_name.get(header.lower())
        if not scheduled:
            errors.append({"cell": f"{letter}3", "message": f"Game {header} is not scheduled for the selected date."})
            continue
        if header.lower() in seen:
            errors.append({"cell": f"{letter}3", "message": f"Duplicate spreadsheet column maps to {header}."})
            continue
        seen[header.lower()] = True
        columns.append({"column": col, "letter": letter, "header": raw_sheet.cell(3, col).value, "game_name": header, "scheduled_id": scheduled.id})
    return columns


def parse_daily_sheet_workbook(uploaded_file, agency, transaction_date):
    warnings = []
    errors = []
    data = validate_workbook_bytes(uploaded_file)
    workbook = load_workbook(data)
    raw_sheet = workbook["ENTER GAME DATA HERE"]
    registration_sheet = workbook["REGISTER SUB-AGENT"]
    if raw_sheet.max_column > MAX_INSPECTED_COLUMNS or registration_sheet.max_column > MAX_INSPECTED_COLUMNS:
        raise ValidationError("Workbook has too many populated columns.")

    workbook_date = parse_workbook_date(raw_sheet["B2"].value)
    if workbook_date is None:
        warnings.append({"cell": "B2", "message": "Workbook date could not be parsed; selected date will be used."})
    elif workbook_date != transaction_date:
        warnings.append({"cell": "B2", "code": "DATE_MISMATCH", "message": f"Workbook date {workbook_date.isoformat()} differs from selected date {transaction_date.isoformat()}."})

    existing_sheet = DailySheet.objects.filter(agency=agency, transaction_date=transaction_date).first()
    scheduled_by_name = scheduled_games_for_date(transaction_date, existing_sheet)
    schedule_snapshot = schedule_snapshot_for_date(transaction_date, existing_sheet)
    game_columns = build_game_columns(raw_sheet, scheduled_by_name, errors)
    if not game_columns:
        errors.append({"cell": "C3:I3", "message": "No spreadsheet game columns matched the selected date schedule."})
    register = registration_lookup(registration_sheet, warnings, errors)
    tpm_matches = {}
    terminal_keys = {item["terminal"].lower() for item in register.values()}
    for code in TPMCode.objects.select_related("person").filter(person__agency=agency, is_active=True, person__is_active=True):
        if code.code.lower() in terminal_keys:
            tpm_matches.setdefault(code.code.lower(), []).append(code)

    rows = []
    seen_sub_codes = set()
    seen_tpm_codes = set()
    ignored_blank_rows = 0
    ignored_zero_rows = 0
    for row_index in TRANSACTION_ROWS:
        sub_cell = raw_sheet.cell(row_index, 2)
        sub_ref = excel_ref(row_index, 2)
        sub_code = normalize_identifier(sub_cell.value, sub_ref, warnings)
        amounts = {}
        row_errors = []
        row_total = money("0")
        for column in game_columns:
            cell = raw_sheet.cell(row_index, column["column"])
            amount = decimal_from_cell(cell, excel_ref(row_index, column["column"]), row_errors)
            amounts[column["game_name"]] = str(amount)
            row_total += amount
        raw_sales_present = any(raw_sheet.cell(row_index, col).value not in (None, "", 0) for col in SALES_COLUMNS)
        if not sub_code and not raw_sales_present:
            ignored_blank_rows += 1
            continue
        if not sub_code and raw_sales_present:
            errors.append({"row": row_index, "cell": sub_ref, "message": "Sales row is missing SUB AGT NOS."})
            continue
        if sub_code in seen_sub_codes:
            errors.append({"row": row_index, "cell": sub_ref, "message": "Duplicate SUB AGT NOS in upload."})
        seen_sub_codes.add(sub_code)
        registration = register.get(sub_code)
        if not registration:
            errors.append({"row": row_index, "cell": sub_ref, "message": "SUB AGT NOS was not found in REGISTER SUB-AGENT."})
            continue
        terminal = registration["terminal"]
        matches = tpm_matches.get(terminal.lower(), [])
        if not matches:
            errors.append({"row": row_index, "cell": sub_ref, "message": f"TERMINAL NOS {terminal} does not match a system TPM code."})
            continue
        if len(matches) > 1:
            errors.append({"row": row_index, "cell": sub_ref, "message": f"TERMINAL NOS {terminal} matches multiple system TPM codes."})
            continue
        tpm = matches[0]
        if tpm.code.lower() in seen_tpm_codes:
            errors.append({"row": row_index, "cell": sub_ref, "message": "Duplicate resolved TPM Code in upload."})
        seen_tpm_codes.add(tpm.code.lower())
        if row_errors:
            errors.extend({"row": row_index, **error} for error in row_errors)
            continue
        if row_total == 0:
            ignored_zero_rows += 1
            warnings.append({"row": row_index, "cell": sub_ref, "message": "Row has a valid identifier but all sales are blank or zero; ignored."})
            continue
        workbook_name = registration["workbook_name"]
        if not workbook_name:
            warnings.append({"row": row_index, "cell": sub_ref, "message": "Workbook name is missing; system person name will be used."})
        elif workbook_name.strip().lower() != tpm.person.full_name.strip().lower():
            warnings.append({"row": row_index, "cell": sub_ref, "message": "Workbook name differs from system person name."})
        rows.append({
            "excel_row": row_index,
            "sub_agent_no": sub_code,
            "tpm_code": tpm.code,
            "tpm_code_id": tpm.id,
            "person_name": tpm.person.full_name,
            "workbook_name": workbook_name,
            "agent_type": tpm.person.agent_type,
            "amounts": amounts,
            "net_sales": str(row_total),
            "to_pay": str(money(row_total * Decimal("0.95"))),
        })

    existing_transaction_count = existing_sheet.transactions.count() if existing_sheet else 0
    payload = {
        "file_name": safe_filename(uploaded_file.name),
        "file_hash": hashlib.sha256(data).hexdigest(),
        "agency": agency.id,
        "agency_name": agency.name,
        "transaction_date": transaction_date.isoformat(),
        "workbook_date": workbook_date.isoformat() if workbook_date else None,
        "valid_row_count": len(rows),
        "ignored_blank_rows": ignored_blank_rows,
        "ignored_zero_rows": ignored_zero_rows,
        "game_columns": game_columns,
        "schedule_snapshot": schedule_snapshot,
        "rows": rows,
        "sheet_totals": {
            "net_sales": str(money(sum(Decimal(row["net_sales"]) for row in rows))),
            "to_pay": str(money(sum(Decimal(row["to_pay"]) for row in rows))),
        },
        "existing_sheet": existing_sheet.id if existing_sheet else None,
        "existing_sheet_status": existing_sheet.status if existing_sheet else None,
        "existing_transaction_count": existing_transaction_count,
        "editable_existing_sheet": bool(existing_sheet and existing_sheet.is_accountant_editable),
        "requires_date_mismatch_ack": bool(workbook_date and workbook_date != transaction_date),
    }
    return WorkbookParseResult(payload=payload, warnings=warnings, errors=errors)
