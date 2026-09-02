from datetime import date, timedelta
from io import BytesIO, StringIO
from zipfile import ZIP_DEFLATED, ZipFile

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import TestCase
from django.utils import timezone
from openpyxl import Workbook
from rest_framework import status
from rest_framework.test import APITestCase

from core.importers import parse_daily_sheet_workbook
from core.models import (
    Agency,
    AgentType,
    AuditAction,
    AuditLog,
    DailySheet,
    DailySheetImportBatch,
    DailySheetImportStatus,
    Person,
    TPMCode,
    TPMDailyTransaction,
    UserAgencyAssignment,
)


User = get_user_model()


def workbook_upload(rows=None, headers=None, register_rows=None, raw_date=date(2026, 8, 27), filename="import.xlsx"):
    workbook = Workbook()
    raw = workbook.active
    raw.title = "ENTER GAME DATA HERE"
    register = workbook.create_sheet("REGISTER SUB-AGENT")
    workbook.create_sheet("MUSA RESULTS")
    workbook.create_sheet("Premier Games")
    workbook.create_sheet("Sheet2")
    raw["B2"] = raw_date
    for offset, header in enumerate(headers or ["F/chance", "diamondq", "Inter", "Bingo", "Peoples"], start=3):
        raw.cell(3, offset).value = header
    for index, row in enumerate(rows or [], start=5):
        raw.cell(index, 2).value = row.get("sub")
        for offset, amount in enumerate(row.get("amounts", []), start=3):
            raw.cell(index, offset).value = amount
    for index, row in enumerate(register_rows or [], start=2):
        register.cell(index, 2).value = row[0]
        register.cell(index, 3).value = row[1]
        register.cell(index, 4).value = row[2]
    raw["C226"] = "=SUM(C5:C224)"
    buffer = BytesIO()
    workbook.save(buffer)
    return SimpleUploadedFile(filename, buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


class DailySheetImportParserTests(TestCase):
    def setUp(self):
        call_command("seed_initial_data", stdout=StringIO())
        self.agency = Agency.objects.create(name="Import Agency", code="import-agency")
        self.person = Person.objects.create(agency=self.agency, full_name="System Name", agent_type=AgentType.SUBAGENT)
        self.tpm = TPMCode.objects.create(person=self.person, code="513670124")

    def parse(self, upload):
        return parse_daily_sheet_workbook(upload, self.agency, date(2026, 8, 27))

    def test_required_sheets_registration_lookup_and_alias_game_mapping(self):
        parsed = self.parse(workbook_upload(
            rows=[{"sub": 469001, "amounts": [10, 20, 30, 40, 50]}],
            register_rows=[(469001, "513670124", "System Name")],
        ))

        self.assertEqual(parsed.errors, [])
        self.assertEqual(parsed.payload["valid_row_count"], 1)
        self.assertEqual(parsed.payload["rows"][0]["tpm_code"], "513670124")
        self.assertEqual([column["game_name"] for column in parsed.payload["game_columns"]], ["Fairchance", "Diamond", "International", "Bingo", "Peoples"])
        self.assertEqual(parsed.payload["rows"][0]["net_sales"], "150.00")
        self.assertEqual(parsed.payload["rows"][0]["to_pay"], "142.50")

    def test_h_i_dynamic_headers_and_blank_optional_column(self):
        parsed = self.parse(workbook_upload(
            headers=["F/chance", "diamondq", "Inter", "Bingo", "Peoples", "", "fortune"],
            rows=[{"sub": 469001, "amounts": [1, 2, 3, 4, 5, None, 6]}],
            register_rows=[(469001, "513670124", "System Name")],
        ))

        self.assertEqual(parsed.errors, [])
        self.assertIn("Fortune", [column["game_name"] for column in parsed.payload["game_columns"]])

    def test_blank_unknown_and_duplicate_game_headers_are_rejected(self):
        blank = self.parse(workbook_upload(headers=["F/chance", ""], rows=[{"sub": 469001, "amounts": [1, 2]}], register_rows=[(469001, "513670124", "System Name")]))
        unknown = self.parse(workbook_upload(headers=["F/chance", "Mystery"], rows=[{"sub": 469001, "amounts": [1, 2]}], register_rows=[(469001, "513670124", "System Name")]))
        duplicate = self.parse(workbook_upload(headers=["F/chance", "Fairchance"], rows=[{"sub": 469001, "amounts": [1, 2]}], register_rows=[(469001, "513670124", "System Name")]))

        self.assertTrue(any("blank game header" in error["message"] for error in blank.errors))
        self.assertTrue(any("not scheduled" in error["message"] for error in unknown.errors))
        self.assertTrue(any("Duplicate" in error["message"] for error in duplicate.errors))

    def test_blank_zero_missing_identifier_unknown_duplicate_and_formula_rows(self):
        upload = workbook_upload(
            rows=[
                {"sub": None, "amounts": [None, None, None]},
                {"sub": 469001, "amounts": [0, None, 0]},
                {"sub": None, "amounts": [5]},
                {"sub": 469002, "amounts": [5]},
                {"sub": 469001, "amounts": [5]},
            ],
            register_rows=[(469001, "513670124", "Different Name")],
        )
        parsed = self.parse(upload)

        self.assertEqual(parsed.payload["ignored_blank_rows"], 216)
        self.assertEqual(parsed.payload["ignored_zero_rows"], 1)
        self.assertTrue(any("missing SUB AGT NOS" in error["message"] for error in parsed.errors))
        self.assertTrue(any("not found" in error["message"] for error in parsed.errors))
        self.assertTrue(any("Duplicate SUB" in error["message"] for error in parsed.errors))
        self.assertTrue(any("Workbook name differs" in warning["message"] for warning in parsed.warnings))
        self.assertTrue(any("lost leading zeroes" in warning["message"] for warning in parsed.warnings))

    def test_invalid_negative_oversized_formula_and_date_mismatch(self):
        upload = workbook_upload(rows=[{"sub": 469001, "amounts": [-1, "text", True, 10**13]}], register_rows=[(469001, "513670124", "System Name")], raw_date=date(2026, 8, 28))
        parsed = self.parse(upload)

        self.assertTrue(parsed.payload["requires_date_mismatch_ack"])
        self.assertTrue(any("negative" in error["message"] for error in parsed.errors))
        self.assertTrue(any("literal number" in error["message"] for error in parsed.errors))
        self.assertTrue(any("too large" in error["message"] for error in parsed.errors))

    def test_unsupported_workbooks_are_rejected(self):
        not_xlsx = SimpleUploadedFile("bad.xlsx", b"not an xlsx")
        oversized = SimpleUploadedFile("big.xlsx", b"PK\x03\x04" + b"0" * (5 * 1024 * 1024 + 1))
        macro = BytesIO()
        with ZipFile(macro, "w", ZIP_DEFLATED) as package:
            package.writestr("xl/vbaProject.bin", b"x")

        for upload in [not_xlsx, oversized, SimpleUploadedFile("macro.xlsx", macro.getvalue())]:
            with self.assertRaises(Exception):
                self.parse(upload)


class DailySheetImportWorkflowTests(APITestCase):
    def setUp(self):
        call_command("seed_initial_data", stdout=StringIO())
        self.admin = User.objects.create_superuser("admin@example.com", "pass-12345", full_name="Admin")
        self.accountant = User.objects.create_user("acct@example.com", "pass-12345", full_name="Acct")
        self.unassigned = User.objects.create_user("other@example.com", "pass-12345", full_name="Other")
        self.agency = Agency.objects.create(name="Import Agency", code="import-agency")
        UserAgencyAssignment.objects.create(user=self.accountant, agency=self.agency, can_create=True)
        self.person = Person.objects.create(agency=self.agency, full_name="System Name", agent_type=AgentType.SUBAGENT)
        self.tpm = TPMCode.objects.create(person=self.person, code="513670124")

    def preview(self, user=None):
        self.client.force_authenticate(user or self.accountant)
        return self.client.post(
            "/api/daily-sheet-imports/preview/",
            {"agency": self.agency.id, "transaction_date": "2026-08-27", "file": workbook_upload(rows=[{"sub": 469001, "amounts": [100, 0, 0, 0, 0]}], register_rows=[(469001, "513670124", "System Name")])},
            format="multipart",
        )

    def test_assigned_accountant_preview_confirm_creates_draft_atomically(self):
        preview = self.preview()
        self.assertEqual(preview.status_code, status.HTTP_201_CREATED)

        confirm = self.client.post(f"/api/daily-sheet-imports/{preview.data['id']}/confirm/", {"replace_existing": False}, format="json")

        self.assertEqual(confirm.status_code, status.HTTP_200_OK)
        sheet = DailySheet.objects.get(pk=confirm.data["daily_sheet"])
        self.assertEqual(sheet.status, "DRAFT")
        self.assertEqual(sheet.transactions.count(), 1)
        txn = sheet.transactions.get()
        self.assertEqual(txn.net_sales, 100)
        self.assertEqual(txn.to_pay, 95)
        self.assertTrue(AuditLog.objects.filter(action=AuditAction.IMPORT_PREVIEWED).exists())
        self.assertTrue(AuditLog.objects.filter(action=AuditAction.IMPORT_CONFIRMED).exists())

    def test_permissions_and_batch_ownership_are_enforced(self):
        denied = self.preview(self.unassigned)
        self.assertEqual(denied.status_code, status.HTTP_403_FORBIDDEN)
        preview = self.preview(self.accountant)
        self.client.force_authenticate(self.unassigned)
        self.assertEqual(self.client.get(f"/api/daily-sheet-imports/{preview.data['id']}/").status_code, status.HTTP_404_NOT_FOUND)

        self.client.force_authenticate(self.admin)
        allowed = self.client.post(
            "/api/daily-sheet-imports/preview/",
            {"agency": self.agency.id, "transaction_date": "2026-08-27", "file": workbook_upload(rows=[{"sub": 469001, "amounts": [100]}], register_rows=[(469001, "513670124", "System Name")])},
            format="multipart",
        )
        self.assertEqual(allowed.status_code, status.HTTP_201_CREATED)

    def test_replacement_requires_flag_and_replaces_not_merges(self):
        sheet = DailySheet.objects.create(agency=self.agency, transaction_date=date(2026, 8, 27), created_by=self.accountant)
        sheet.copy_weekday_games()
        TPMDailyTransaction.objects.create(daily_sheet=sheet, tpm_code=self.tpm, created_by=self.accountant, updated_by=self.accountant)
        preview = self.preview()

        blocked = self.client.post(f"/api/daily-sheet-imports/{preview.data['id']}/confirm/", {"replace_existing": False}, format="json")
        confirmed = self.client.post(f"/api/daily-sheet-imports/{preview.data['id']}/confirm/", {"replace_existing": True}, format="json")

        self.assertEqual(blocked.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(confirmed.status_code, status.HTTP_200_OK)
        self.assertEqual(sheet.transactions.count(), 1)

    def test_submitted_sheet_cancelled_expired_and_double_confirm_blocked(self):
        preview = self.preview()
        batch = DailySheetImportBatch.objects.get(pk=preview.data["id"])
        batch.expires_at = timezone.now() - timedelta(minutes=1)
        batch.save(update_fields=["expires_at"])
        expired = self.client.post(f"/api/daily-sheet-imports/{batch.id}/confirm/", {}, format="json")
        self.assertEqual(expired.status_code, status.HTTP_400_BAD_REQUEST)

        preview = self.preview()
        cancelled = self.client.post(f"/api/daily-sheet-imports/{preview.data['id']}/cancel/", {}, format="json")
        self.assertEqual(cancelled.status_code, status.HTTP_200_OK)
        self.assertEqual(self.client.post(f"/api/daily-sheet-imports/{preview.data['id']}/confirm/", {}, format="json").status_code, status.HTTP_400_BAD_REQUEST)

        preview = self.preview()
        first = self.client.post(f"/api/daily-sheet-imports/{preview.data['id']}/confirm/", {}, format="json")
        second = self.client.post(f"/api/daily-sheet-imports/{preview.data['id']}/confirm/", {}, format="json")
        self.assertEqual(first.status_code, status.HTTP_200_OK)
        self.assertEqual(second.status_code, status.HTTP_400_BAD_REQUEST)

    def test_submitted_sheet_and_target_changes_are_blocked(self):
        sheet = DailySheet.objects.create(agency=self.agency, transaction_date=date(2026, 8, 27), created_by=self.accountant, status="SUBMITTED")
        sheet.copy_weekday_games()
        preview = self.preview()
        self.assertEqual(self.client.post(f"/api/daily-sheet-imports/{preview.data['id']}/confirm/", {"replace_existing": True}, format="json").status_code, status.HTTP_400_BAD_REQUEST)

    def test_date_ack_required_and_row_contents_not_audit_logged(self):
        self.client.force_authenticate(self.accountant)
        preview = self.client.post(
            "/api/daily-sheet-imports/preview/",
            {"agency": self.agency.id, "transaction_date": "2026-08-27", "file": workbook_upload(rows=[{"sub": 469001, "amounts": [100]}], register_rows=[(469001, "513670124", "System Name")], raw_date=date(2026, 8, 28))},
            format="multipart",
        )
        blocked = self.client.post(f"/api/daily-sheet-imports/{preview.data['id']}/confirm/", {}, format="json")
        confirmed = self.client.post(f"/api/daily-sheet-imports/{preview.data['id']}/confirm/", {"acknowledge_date_mismatch": True}, format="json")

        self.assertEqual(blocked.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(confirmed.status_code, status.HTTP_200_OK)
        audit = AuditLog.objects.filter(action=AuditAction.IMPORT_PREVIEWED).latest("created_at")
        self.assertNotIn("rows", audit.new_values)
