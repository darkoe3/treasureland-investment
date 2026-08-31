from datetime import date, time
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.db import connection
from django.test import TestCase
from django.test.utils import CaptureQueriesContext
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from core.models import (
    Agency,
    AgentType,
    AuditLog,
    DailySheet,
    DailySheetGame,
    DailySheetStatus,
    Game,
    OmittedTerminal,
    Person,
    TPMCode,
    TPMDailyTransaction,
    TransactionGameSale,
    UserAgencyAssignment,
    WeeklyGameSchedule,
)
from core.reports import build_report, resolve_report_query


User = get_user_model()


class Phase5Mixin:
    def setUpBase(self):
        self.admin = User.objects.create_superuser("admin@example.com", "AdminPass123!", full_name="Admin User")
        self.acct = User.objects.create_user("acct@example.com", "AcctPass123!", full_name="Acct User")
        self.agencies = [
            Agency.objects.create(name="Musa 1", code="musa-1"),
            Agency.objects.create(name="Musa 2", code="musa-2"),
            Agency.objects.create(name="Omolade", code="omolade"),
            Agency.objects.create(name="Treasure Land", code="treasure-land"),
            Agency.objects.create(name="Sango", code="sango"),
        ]
        self.agency = self.agencies[0]
        UserAgencyAssignment.objects.create(user=self.acct, agency=self.agency, can_export=True, can_create=True, can_edit=True)
        self.person = Person.objects.create(agency=self.agency, full_name="Ayo", agent_type=AgentType.MAIN_AGENT)
        self.sub = Person.objects.create(agency=self.agency, full_name="=Formula Person", agent_type=AgentType.SUBAGENT)
        self.tpm_one = TPMCode.objects.create(person=self.person, code="+TPM-A")
        self.tpm_two = TPMCode.objects.create(person=self.person, code="TPM-B")
        self.tpm_omitted = TPMCode.objects.create(person=self.person, code="TPM-OMIT")
        self.sub_tpm = TPMCode.objects.create(person=self.sub, code="@SUB-A")
        self.royal = Game.objects.create(name="Royal")
        self.international = Game.objects.create(name="INTERNATIONAL")
        self.mark = Game.objects.create(name="Mark II")
        for weekday in range(1, 8):
            WeeklyGameSchedule.objects.create(game=self.royal, weekday=weekday, closing_time=time(10), draw_time=time(11), display_order=1)
            WeeklyGameSchedule.objects.create(game=self.international, weekday=weekday, closing_time=time(12), draw_time=time(13), display_order=2)

    def sheet(self, day, status_value=DailySheetStatus.APPROVED, incoming="285.00", tax="12.00"):
        sheet = DailySheet.objects.create(
            agency=self.agency,
            transaction_date=day,
            incoming_funds=Decimal(incoming),
            tax=Decimal(tax),
            created_by=self.acct,
        )
        sheet.copy_weekday_games()
        sheet.status = status_value
        sheet.save()
        return sheet

    def txn(self, sheet, tpm, amounts):
        txn = TPMDailyTransaction.objects.create(daily_sheet=sheet, tpm_code=tpm, created_by=self.acct, updated_by=self.acct)
        games = list(sheet.sheet_games.order_by("display_order", "id"))
        for game, amount in zip(games, amounts):
            TransactionGameSale.objects.create(transaction=txn, daily_sheet_game=game, amount=Decimal(amount))
        return txn


class Phase5ReportServiceTests(Phase5Mixin, TestCase):
    def setUp(self):
        self.setUpBase()

    def test_period_resolution_daily_weekly_monthly_custom_and_validation(self):
        daily = resolve_report_query({"agency": str(self.agency.id), "period": "daily", "date": "2026-08-24"})
        weekly = resolve_report_query({"agency": str(self.agency.id), "period": "weekly", "date": "2026-08-27"})
        monthly = resolve_report_query({"agency": str(self.agency.id), "period": "monthly", "month": "2", "year": "2024"})
        custom = resolve_report_query({"agency": str(self.agency.id), "period": "custom", "start_date": "2026-08-24", "end_date": "2026-08-31"})
        self.assertEqual((daily["start_date"], daily["end_date"]), (date(2026, 8, 24), date(2026, 8, 24)))
        self.assertEqual((weekly["start_date"], weekly["end_date"]), (date(2026, 8, 24), date(2026, 8, 30)))
        self.assertEqual((monthly["start_date"], monthly["end_date"]), (date(2024, 2, 1), date(2024, 2, 29)))
        self.assertEqual((custom["start_date"], custom["end_date"]), (date(2026, 8, 24), date(2026, 8, 31)))
        with self.assertRaises(Exception):
            resolve_report_query({"agency": str(self.agency.id), "period": "custom", "start_date": "2026-08-31", "end_date": "2026-08-24"})

    def test_empty_report_and_all_five_agencies_are_supported(self):
        for agency in self.agencies:
            report = build_report({"agency": str(agency.id), "period": "daily", "date": "2026-08-24"}, self.admin)
            self.assertEqual(report["header"]["agency"], agency.name)
            self.assertEqual(report["summary"]["daily_sheet_count"], 0)

    def test_dynamic_game_union_multi_tpm_person_totals_and_calculations(self):
        first = self.sheet(date(2026, 8, 24), incoming="237.50", tax="99.99")
        self.txn(first, self.tpm_one, ["100.00", "0.00"])
        self.txn(first, self.tpm_two, ["50.00", "0.00"])
        self.txn(first, self.sub_tpm, ["100.00", "0.00"])
        OmittedTerminal.objects.create(daily_sheet=first, tpm_code=self.tpm_omitted, reason="Closed later", recorded_by=self.acct)

        second = self.sheet(date(2026, 8, 25), incoming="95.00", tax="0.00")
        extra = DailySheetGame.objects.create(
            daily_sheet=second,
            game=self.mark,
            game_name_snapshot="Mark II",
            closing_time_snapshot=time(14),
            draw_time_snapshot=time(15),
            display_order=3,
        )
        txn = self.txn(second, self.tpm_one, ["100.005", "0.00"])
        TransactionGameSale.objects.create(transaction=txn, daily_sheet_game=extra, amount=Decimal("25.00"))

        report = build_report({"agency": str(self.agency.id), "period": "weekly", "date": "2026-08-27"}, self.admin)
        self.assertEqual([game["name"] for game in report["game_columns"]], ["Royal", "INTERNATIONAL", "Mark II"])
        self.assertEqual(report["summary"]["daily_sheet_count"], 2)
        self.assertEqual(report["summary"]["transaction_row_count"], 4)
        self.assertEqual(report["summary"]["distinct_people_count"], 2)
        self.assertEqual(report["summary"]["distinct_tpm_code_count"], 3)
        self.assertEqual(report["summary"]["total_net_sales"], Decimal("375.01"))
        self.assertEqual(report["summary"]["total_commission"], Decimal("18.75"))
        self.assertEqual(report["summary"]["total_to_pay"], Decimal("356.26"))
        self.assertEqual(report["summary"]["total_subagent_share"], Decimal("2.00"))
        self.assertEqual(report["summary"]["total_organisation_share"], Decimal("3.00"))
        self.assertEqual(report["summary"]["total_manual_tax"], Decimal("99.99"))
        self.assertEqual(report["summary"]["total_actual_amount_received"], Decimal("332.50"))
        self.assertEqual(report["summary"]["total_difference"], Decimal("-23.76"))
        ayo_rows = [row for row in report["details"] if row["name"] == "Ayo"]
        self.assertEqual(ayo_rows[0]["total"], Decimal("261.26"))
        self.assertEqual(ayo_rows[1]["total"], "")
        self.assertEqual(sum(row["to_pay"] for row in report["details"]), report["summary"]["total_to_pay"])

    def test_report_builder_uses_bounded_query_count_for_multiple_sheets(self):
        for offset in range(3):
            sheet = self.sheet(date(2026, 8, 24 + offset), incoming="95.00", tax="0.00")
            self.txn(sheet, self.tpm_one, ["100.00", "0.00"])
        with CaptureQueriesContext(connection) as queries:
            report = build_report({"agency": str(self.agency.id), "period": "custom", "start_date": "2026-08-24", "end_date": "2026-08-26"}, self.admin)
        self.assertEqual(report["summary"]["daily_sheet_count"], 3)
        self.assertLessEqual(len(queries), 8)


class Phase5ReportAPITests(Phase5Mixin, APITestCase):
    def setUp(self):
        self.setUpBase()

    def test_super_admin_access_and_accountant_denial_even_with_export_permission(self):
        self.client.force_authenticate(self.admin)
        ok = self.client.get(f"/api/reports/agency-summary/?agency={self.agency.id}&period=daily&date=2026-08-24")
        self.assertEqual(ok.status_code, status.HTTP_200_OK)
        self.client.force_authenticate(self.acct)
        denied = self.client.get(f"/api/reports/agency-summary/?agency={self.agency.id}&period=daily&date=2026-08-24")
        denied_export = self.client.get(f"/api/reports/agency-summary/export/?agency={self.agency.id}&period=daily&date=2026-08-24")
        self.assertEqual(denied.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(denied_export.status_code, status.HTTP_403_FORBIDDEN)

    def test_default_approved_status_and_explicit_operational_statuses(self):
        approved = self.sheet(date(2026, 8, 24), DailySheetStatus.APPROVED)
        draft = self.sheet(date(2026, 8, 25), DailySheetStatus.DRAFT)
        self.txn(approved, self.tpm_one, ["100.00", "0.00"])
        self.txn(draft, self.tpm_two, ["200.00", "0.00"])
        self.client.force_authenticate(self.admin)
        default = self.client.get(f"/api/reports/agency-summary/?agency={self.agency.id}&period=custom&start_date=2026-08-24&end_date=2026-08-25")
        explicit = self.client.get(f"/api/reports/agency-summary/?agency={self.agency.id}&period=custom&start_date=2026-08-24&end_date=2026-08-25&status=APPROVED&status=DRAFT")
        self.assertEqual(default.data["summary"]["daily_sheet_count"], 1)
        self.assertEqual(default.data["header"]["selected_statuses"], ["APPROVED"])
        self.assertEqual(explicit.data["summary"]["daily_sheet_count"], 2)
        self.assertFalse(explicit.data["header"]["is_final"])
        self.assertEqual(explicit.data["header"]["label"], "Operational non-final report")

    def test_export_workbook_headers_numeric_values_formula_safety_and_audit(self):
        sheet = self.sheet(date(2026, 8, 24), incoming="95.00", tax="0.00")
        self.txn(sheet, self.tpm_one, ["100.00", "0.00"])
        self.client.force_authenticate(self.admin)
        response = self.client.get(f"/api/reports/agency-summary/export/?agency={self.agency.id}&period=daily&date=2026-08-24")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn("treasureland-musa-1-daily-2026-08-24-to-2026-08-24.xlsx", response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        ws = workbook.active
        self.assertEqual(ws["A1"].value, "Treasureland Investment Limited")
        values = [[cell.value for cell in row] for row in ws.iter_rows()]
        flat = [item for row in values for item in row]
        self.assertIn("Daily reconciliation", flat)
        self.assertIn("TPM Code", flat)
        self.assertIn("'+TPM-A", flat)
        self.assertTrue(any(isinstance(item, (int, float)) and item == 100 for item in flat))
        self.assertTrue(AuditLog.objects.filter(action="REPORT_EXPORTED", agency=self.agency).exists())

    def test_json_and_excel_totals_agree(self):
        sheet = self.sheet(date(2026, 8, 24), incoming="95.00", tax="0.00")
        self.txn(sheet, self.tpm_one, ["100.00", "0.00"])
        self.client.force_authenticate(self.admin)
        preview = self.client.get(f"/api/reports/agency-summary/?agency={self.agency.id}&period=daily&date=2026-08-24")
        export = self.client.get(f"/api/reports/agency-summary/export/?agency={self.agency.id}&period=daily&date=2026-08-24")
        ws = load_workbook(BytesIO(export.content), data_only=True).active
        flat = [cell.value for row in ws.iter_rows() for cell in row]
        self.assertIn(float(Decimal(preview.data["summary"]["total_net_sales"])), flat)
        self.assertIn(float(Decimal(preview.data["summary"]["total_to_pay"])), flat)
