import re
from calendar import monthrange
from collections import OrderedDict, defaultdict
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from zoneinfo import ZoneInfo

from django.db.models import Prefetch
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.exceptions import NotFound, ValidationError

from .models import (
    AgentType,
    Agency,
    AuditAction,
    AuditLog,
    COMMISSION_RATE,
    DailySheet,
    DailySheetStatus,
    OmittedTerminal,
    ORGANISATION_SUBAGENT_RATE,
    SUBAGENT_RATE,
    TPMDailyTransaction,
    TO_PAY_RATE,
    json_safe_value,
    money,
)


ORG_NAME = "Treasureland Investment Limited"
ACCRA_ZONE = ZoneInfo("Africa/Accra")
STATUS_ORDER = [choice.value for choice in DailySheetStatus]
STATUS_VALUES = set(STATUS_ORDER)
STATUS_LABELS = {choice.value: choice.label for choice in DailySheetStatus}


def log_report_audit(user, agency, action, object_id, values, description):
    AuditLog.objects.create(
        user=user,
        agency=agency,
        action=action,
        model_name="Report",
        object_id=str(object_id),
        new_values={key: json_safe_value(value) for key, value in values.items()},
        description=description,
    )


def parse_iso_date(value, field):
    if not value:
        raise ValidationError({field: f"{field} is required."})
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise ValidationError({field: "Use YYYY-MM-DD."}) from exc


def resolve_report_query(params):
    agency_id = params.get("agency")
    if not agency_id:
        raise ValidationError({"agency": "agency is required."})
    if not str(agency_id).isdigit():
        raise ValidationError({"agency": "agency must be a numeric active agency ID."})
    try:
        agency = Agency.objects.get(pk=agency_id, is_active=True)
    except Agency.DoesNotExist as exc:
        raise NotFound("Agency was not found.")

    period = (params.get("period") or "").lower()
    if period not in {"daily", "weekly", "monthly", "custom"}:
        raise ValidationError({"period": "period must be daily, weekly, monthly or custom."})

    if period == "daily":
        selected = parse_iso_date(params.get("date"), "date")
        start_date = end_date = selected
    elif period == "weekly":
        selected = parse_iso_date(params.get("date"), "date")
        start_date = selected - timedelta(days=selected.weekday())
        end_date = start_date + timedelta(days=6)
    elif period == "monthly":
        month_value = params.get("month")
        year_value = params.get("year")
        if not month_value:
            raise ValidationError({"month": "month is required."})
        if not year_value:
            raise ValidationError({"year": "year is required."})
        if not str(month_value).isdigit() or not 1 <= int(month_value) <= 12:
            raise ValidationError({"month": "month must be between 1 and 12."})
        if not str(year_value).isdigit():
            raise ValidationError({"year": "year must be numeric."})
        month = int(month_value)
        year = int(year_value)
        start_date = date(year, month, 1)
        end_date = date(year, month, monthrange(year, month)[1])
    else:
        start_date = parse_iso_date(params.get("start_date"), "start_date")
        end_date = parse_iso_date(params.get("end_date"), "end_date")
        if start_date > end_date:
            raise ValidationError({"end_date": "end_date must be on or after start_date."})

    raw_statuses = params.getlist("status") if hasattr(params, "getlist") else []
    if not raw_statuses:
        status_text = params.get("statuses") or params.get("status") or ""
        raw_statuses = [part for part in re.split(r"[, ]+", status_text) if part]
    statuses = [item.upper() for item in raw_statuses] or [DailySheetStatus.APPROVED]
    invalid = sorted(set(statuses) - STATUS_VALUES)
    if invalid:
        raise ValidationError({"status": f"Unsupported status: {', '.join(invalid)}."})
    statuses = sorted(set(statuses), key=STATUS_ORDER.index)

    return {
        "agency": agency,
        "period": period,
        "start_date": start_date,
        "end_date": end_date,
        "statuses": statuses,
        "is_final": statuses == [DailySheetStatus.APPROVED],
    }


def game_key(sheet_game):
    if sheet_game.game_id:
        return f"game:{sheet_game.game_id}"
    normalized = re.sub(r"\s+", " ", sheet_game.game_name_snapshot.strip().casefold())
    return f"name:{normalized}"


def build_report(params, user, audit_action=AuditAction.REPORT_PREVIEWED):
    query = resolve_report_query(params)
    txns = TPMDailyTransaction.objects.select_related("tpm_code", "tpm_code__person").prefetch_related(
        "sales", "sales__daily_sheet_game"
    )
    sheets = list(
        DailySheet.objects.filter(
            agency=query["agency"],
            transaction_date__gte=query["start_date"],
            transaction_date__lte=query["end_date"],
            status__in=query["statuses"],
        )
        .select_related("agency")
        .prefetch_related(
            "sheet_games",
            Prefetch("transactions", queryset=txns),
            Prefetch("omitted_terminals", queryset=OmittedTerminal.objects.filter(is_active=True)),
        )
        .order_by("transaction_date", "id")
    )

    game_columns = OrderedDict()
    for sheet in sheets:
        for sheet_game in sheet.sheet_games.all():
            key = game_key(sheet_game)
            if key not in game_columns:
                game_columns[key] = {
                    "key": key,
                    "name": sheet_game.game_name_snapshot,
                    "display_order": sheet_game.display_order,
                    "first_date": sheet.transaction_date.isoformat(),
                }
    game_columns = OrderedDict(sorted(game_columns.items(), key=lambda item: (item[1]["display_order"], item[1]["name"].casefold(), item[1]["key"])))

    detail_map = {}
    person_totals = defaultdict(lambda: money("0"))
    daily_rows = []
    transaction_count = 0
    omitted_total = 0
    total_tax = money("0")
    total_received = money("0")
    total_net = money("0")
    total_to_pay = money("0")
    total_commission = money("0")
    total_subagent_share = money("0")
    total_org_share = money("0")

    for sheet in sheets:
        txns_for_sheet = list(sheet.transactions.all())
        sheet_net = money("0")
        sheet_subagent_net = money("0")
        for txn in txns_for_sheet:
            txn_net = money(sum((sale.amount for sale in txn.sales.all()), Decimal("0")))
            sheet_net = money(sheet_net + txn_net)
            if txn.agent_type_snapshot == AgentType.SUBAGENT:
                sheet_subagent_net = money(sheet_subagent_net + txn_net)
        sheet_to_pay = money(sheet_net * TO_PAY_RATE)
        sheet_commission = money(sheet_net * COMMISSION_RATE)
        sheet_tax = money(sheet.tax)
        sheet_received = money(sheet.incoming_funds)
        sheet_difference = money(sheet_received - sheet_to_pay)
        omitted_count = len(list(sheet.omitted_terminals.all()))
        transaction_count += len(txns_for_sheet)
        omitted_total += omitted_count
        total_tax = money(total_tax + sheet_tax)
        total_received = money(total_received + sheet_received)
        total_net = money(total_net + sheet_net)
        total_to_pay = money(total_to_pay + sheet_to_pay)
        total_commission = money(total_commission + sheet_commission)
        total_subagent_share = money(total_subagent_share + money(sheet_subagent_net * SUBAGENT_RATE))
        total_org_share = money(total_org_share + money(sheet_subagent_net * ORGANISATION_SUBAGENT_RATE))
        daily_rows.append({
            "date": sheet.transaction_date.isoformat(),
            "status": sheet.status,
            "net_sales": sheet_net,
            "commission": sheet_commission,
            "to_pay": sheet_to_pay,
            "tax": sheet_tax,
            "actual_amount_received": sheet_received,
            "difference": sheet_difference,
            "transaction_count": len(txns_for_sheet),
            "omitted_terminal_count": omitted_count,
        })

        for txn in txns_for_sheet:
            person_id = txn.tpm_code.person_id
            detail_key = (person_id, txn.tpm_code_id)
            row = detail_map.setdefault(detail_key, {
                "person": person_id,
                "name": txn.person_name_snapshot,
                "tpm_code": txn.tpm_code.code,
                "games": {key: money("0") for key in game_columns},
                "net_sales": money("0"),
                "to_pay": money("0"),
                "total": "",
                "agent_type": txn.agent_type_snapshot,
            })
            row_net = money("0")
            for sale in txn.sales.all():
                key = game_key(sale.daily_sheet_game)
                row["games"][key] = money(row["games"].get(key, money("0")) + sale.amount)
                row_net = money(row_net + sale.amount)
            row_to_pay = money(row_net * Decimal("0.95"))
            row["net_sales"] = money(row["net_sales"] + row_net)
            row["to_pay"] = money(row["to_pay"] + row_to_pay)
            person_totals[person_id] = money(person_totals[person_id] + row_to_pay)

    detail_rows = sorted(detail_map.values(), key=lambda row: (row["name"].casefold(), row["person"], row["tpm_code"].casefold()))
    seen_people = set()
    for index, row in enumerate(detail_rows, start=1):
        row["no"] = index
        row["person_total"] = person_totals[row["person"]] if row["person"] not in seen_people else ""
        row["total"] = row["person_total"]
        seen_people.add(row["person"])

    report = {
        "header": {
            "organisation": ORG_NAME,
            "agency": query["agency"].name,
            "report_type": query["period"],
            "start_date": query["start_date"].isoformat(),
            "end_date": query["end_date"].isoformat(),
            "selected_statuses": query["statuses"],
            "selected_status_labels": [STATUS_LABELS[item] for item in query["statuses"]],
            "generated_at": timezone.localtime(timezone.now(), ACCRA_ZONE).isoformat(),
            "generated_by": user.full_name or user.email,
            "is_final": query["is_final"],
            "label": "Official approved report" if query["is_final"] else "Operational non-final report",
        },
        "summary": {
            "daily_sheet_count": len(sheets),
            "transaction_row_count": transaction_count,
            "distinct_people_count": len(person_totals),
            "distinct_tpm_code_count": len(detail_rows),
            "total_net_sales": total_net,
            "total_commission": total_commission,
            "total_to_pay": total_to_pay,
            "total_subagent_share": total_subagent_share,
            "total_organisation_share": total_org_share,
            "total_manual_tax": total_tax,
            "total_actual_amount_received": total_received,
            "total_difference": money(total_received - total_to_pay),
            "total_omitted_terminals": omitted_total,
        },
        "game_columns": list(game_columns.values()),
        "daily_reconciliation": daily_rows,
        "details": detail_rows,
    }
    log_report_audit(user, query["agency"], audit_action, f"{query['agency'].id}:{query['period']}:{query['start_date']}:{query['end_date']}", {
        "agency": query["agency"].name,
        "period": query["period"],
        "start_date": query["start_date"],
        "end_date": query["end_date"],
        "statuses": query["statuses"],
    }, f"{report['header']['label']} generated.")
    return report


def serialize_money(value):
    if isinstance(value, Decimal):
        return str(money(value))
    return value


def serialize_report(report):
    def convert(value):
        if isinstance(value, dict):
            return {key: convert(item) for key, item in value.items()}
        if isinstance(value, list):
            return [convert(item) for item in value]
        return serialize_money(value)
    return convert(report)


def safe_excel_text(value):
    text = "" if value is None else str(value)
    return f"'{text}" if text[:1] in {"=", "+", "-", "@"} else text


def safe_slug(value):
    slug = re.sub(r"[^a-z0-9]+", "-", str(value).lower()).strip("-")
    return slug[:60] or "report"


def safe_sheet_title(value):
    title = re.sub(r"[\[\]:*?/\\]", " ", str(value)).strip()[:31]
    return title or "Report"


def append_kv(ws, key, value):
    ws.append([key, value])
    ws.cell(ws.max_row, 1).font = Font(bold=True)


def build_workbook(report):
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title(report["header"]["agency"])
    blue = "061A63"
    gold = "E9A400"
    header_fill = PatternFill("solid", fgColor=blue)
    gold_fill = PatternFill("solid", fgColor=gold)

    ws["A1"] = ORG_NAME
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(8, 7 + len(report["game_columns"])))

    for key, value in [
        ("Agency", report["header"]["agency"]),
        ("Report type", report["header"]["report_type"].title()),
        ("Date range", f"{report['header']['start_date']} to {report['header']['end_date']}"),
        ("Selected statuses", ", ".join(report["header"]["selected_status_labels"])),
        ("Report label", report["header"]["label"]),
        ("Generated", report["header"]["generated_at"]),
        ("Generated by", report["header"]["generated_by"]),
    ]:
        append_kv(ws, key, safe_excel_text(value))

    ws.append([])
    ws.append(["Summary metric", "Value"])
    ws.cell(ws.max_row, 1).fill = gold_fill
    ws.cell(ws.max_row, 2).fill = gold_fill
    for key, value in report["summary"].items():
        ws.append([key.replace("_", " ").title(), float(value) if isinstance(value, Decimal) else value])

    ws.append([])
    ws.append(["Daily reconciliation"])
    daily_header_row = ws.max_row + 1
    daily_headers = ["Date", "Sheet status", "NET Sales", "Commission", "To Pay", "Tax", "Actual amount received", "Difference", "Transaction count", "Omitted-terminal count"]
    ws.append(daily_headers)
    for cell in ws[daily_header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row in report["daily_reconciliation"]:
        ws.append([row["date"], row["status"], float(row["net_sales"]), float(row["commission"]), float(row["to_pay"]), float(row["tax"]), float(row["actual_amount_received"]), float(row["difference"]), row["transaction_count"], row["omitted_terminal_count"]])

    ws.append([])
    detail_header_row = ws.max_row + 1
    detail_headers = ["No", "Name", "TPM Code", *[game["name"] for game in report["game_columns"]], "NET Sales", "To Pay", "Total"]
    ws.append(detail_headers)
    for cell in ws[detail_header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)
    for row in report["details"]:
        ws.append([
            row["no"],
            safe_excel_text(row["name"]),
            safe_excel_text(row["tpm_code"]),
            *[float(row["games"].get(game["key"], money("0"))) for game in report["game_columns"]],
            float(row["net_sales"]),
            float(row["to_pay"]),
            float(row["total"]) if row["total"] != "" else None,
        ])
    ws.append(["", "Totals", "", *[None for _ in report["game_columns"]], float(report["summary"]["total_net_sales"]), float(report["summary"]["total_to_pay"]), float(report["summary"]["total_to_pay"])])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = '#,##0.00'
    for col_idx, column in enumerate(ws.columns, start=1):
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 32)
    ws.freeze_panes = "A12"
    if report["details"]:
        ws.auto_filter.ref = f"A{detail_header_row}:{get_column_letter(len(detail_headers))}{ws.max_row}"
    return wb


def workbook_response(report):
    buffer = BytesIO()
    build_workbook(report).save(buffer)
    buffer.seek(0)
    filename = f"treasureland-{safe_slug(report['header']['agency'])}-{report['header']['report_type']}-{report['header']['start_date']}-to-{report['header']['end_date']}.xlsx"
    response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
